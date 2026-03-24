"""
Brooke Reams - breams@esri.com
Oct. 16, 2025

Description:
Loops through all feature classes and tables at the root
and feature dataset level and reports basic properties
on each dataset.

ArcGIS Pro 3.5.2
Python 3.11.11

Updates:
11/4/2025:      Fix for tables; changed MakeFeatureLayer to MakeTableView.
12/1/2025:      Sorted feature classes and tables alphabetically.
12/1/2025:      Added "Subtype Code" to output Excel spreadsheet.
3/12/2026:      Enhancement to include Asset Type codes, names, and counts.
3/17/2026:      Updated code to use pandas to get counts from dataframe - this
                update significantly reduced script run-time in the test case
                from 1 hr 45 min to under 30 seconds.
3/17/2026:      Added code to handle invalid and null subtypes, as well as
                invalid asset types.
3/24/2026:      Fix for case of dataframe column names by standardizing to
                upper case.

"""

import arcpy
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pandas as pd

# Overwrite existing output
arcpy.env.overwriteOutput = 1


def log_it(message):
    print(message)
    arcpy.AddMessage(message)


def autofit_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(
            col[0].column
        )  # Get column letter from the first cell in the column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (
                TypeError
            ):  # Handle cases where cell.value might be None or not easily convertible to string
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column].width = adjusted_width


# Tool inputs
in_ws = arcpy.GetParameterAsText(0)
out_xls = arcpy.GetParameterAsText(1)
include_assettypes = arcpy.GetParameter(2)

# Set null integer value
null_int_val = -9999

# Create new workbook and define header
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "Feature Dataset"
ws["B1"] = "Feature Class/Table"
ws["C1"] = "Shape Type"
ws["D1"] = "Record Count"
ws["E1"] = "Subtype Code"
ws["F1"] = "Subtype Name"
ws["G1"] = "Subtype Count"
if include_assettypes:
    ws["H1"] = "Asset Type Code"
    ws["I1"] = "Asset Type Name"
    ws["J1"] = "Asset Type Count"


# Set workspace environment
arcpy.env.workspace = in_ws

# Initialize list to store data properties
records = []

# Loop through all feature classes/tables in gdb and in all feature datasets
fds_list = arcpy.ListDatasets(feature_type="Feature")
fds_list.sort()
fds_list.append("")
for fds in fds_list:
    # Get feature classes/tables
    if fds == "":
        log_it(f"Processing stand-alone datasets")
        fds = "<standalone>"
        fc_list = arcpy.ListFeatureClasses()
        tbl_list = arcpy.ListTables()
        fc_list.sort()
        tbl_list.sort()
        ds_list = fc_list + tbl_list
    else:
        log_it(f"Processing feature dataset: {fds}")
        ds_list = arcpy.ListFeatureClasses(feature_dataset=fds)
        ds_list.sort()

    for ds in ds_list:
        log_it(f"Processing dataset: {ds}")

        # Describe dataset to get properties
        desc = arcpy.Describe(ds)
        # Get shape type
        try:
            shape_type = desc.shapeType
        except:
            shape_type = ""
            pass

        # Get subtype info
        subtype_fld = desc.subtypeFieldName

        # Convert to pandas dataframe
        if subtype_fld:
            flds_list = [
                fld.name
                for fld in arcpy.ListFields(ds)
                if fld.name.upper() in (subtype_fld.upper(), "ASSETTYPE")
            ]
        else:
            flds_list = [
                fld.name for fld in arcpy.ListFields(ds) if fld.type == "String"
            ]
        numpy_array = arcpy.da.FeatureClassToNumPyArray(
            ds, flds_list, null_value=null_int_val
        )
        df = pd.DataFrame(numpy_array)
        # Standardize column case to upper case
        df.columns = df.columns.str.upper()

        # Get record count
        record_count = len(df)

        subtype_list = []
        sorted_subtype_list = []
        if subtype_fld:
            subtype_dict = arcpy.da.ListSubtypes(ds)
            for subtype_code, subtype_prop in subtype_dict.items():
                subtype_name = subtype_prop["Name"]
                # Get count of records for subtype
                subtype_count = len(df[df[subtype_fld.upper()] == subtype_code])
                if not include_assettypes:
                    subtype_list.append((subtype_code, subtype_name, subtype_count))
                else:
                    assettype_list = []
                    if "ASSETTYPE" in subtype_prop["FieldValues"].keys():
                        domain = subtype_prop["FieldValues"]["ASSETTYPE"][1]
                        if domain.domainType == "CodedValue":
                            for at_code, at_name in domain.codedValues.items():
                                # Get count of records for subtype
                                at_count = len(
                                    df[
                                        (df[subtype_fld.upper()] == subtype_code)
                                        & (df["ASSETTYPE"] == at_code)
                                    ]
                                )
                                assettype_list.append((at_code, at_name, at_count))

                            # Get invalid asset types
                            all_at = list(domain.codedValues.keys())
                            invalid_df = df[
                                (df[subtype_fld.upper()] == subtype_code)
                                & (~df["ASSETTYPE"].isin(all_at))
                            ]
                            invalid_dict = (
                                invalid_df["ASSETTYPE"].value_counts().to_dict()
                            )
                            for at_code, at_count in invalid_dict.items():
                                at_name = "<Invalid Value>"
                                assettype_list.append((at_code, at_name, at_count))
                    subtype_list.append(
                        (subtype_code, subtype_name, subtype_count, assettype_list)
                    )

            # Get record count of values that are not subtypes
            all_subtypes = list(subtype_dict.keys())
            invalid_df = df[~df[subtype_fld.upper()].isin(all_subtypes)]
            invalid_dict = invalid_df[subtype_fld.upper()].value_counts().to_dict()
            for subtype_code, subtype_count in invalid_dict.items():
                subtype_name = "<Invalid Value>"
                if subtype_code == null_int_val:
                    subtype_code = 99999999  # Need to make value number so results can be sorted on subtype code
                    subtype_name = "<Null>"
                if not include_assettypes:
                    subtype_list.append((subtype_code, subtype_name, subtype_count))
                else:
                    assettype_list = []
                    sorted_assettype_list = []
                    if "ASSETTYPE" in subtype_prop["FieldValues"].keys():
                        domain = subtype_prop["FieldValues"]["ASSETTYPE"][1]
                        if domain.domainType == "CodedValue":
                            for at_code, at_name in domain.codedValues.items():
                                # Get count of records for subtype
                                at_count = len(
                                    df[
                                        (df[subtype_fld.upper()] == subtype_code)
                                        & (df["ASSETTYPE"] == at_code)
                                    ]
                                )
                                assettype_list.append((at_code, at_name, at_count))

                            # Get invalid asset types
                            all_at = list(domain.codedValues.keys())
                            invalid_df = df[
                                (df[subtype_fld.upper()] == subtype_code)
                                & (~df["ASSETTYPE"].isin(all_at))
                            ]
                            invalid_dict = (
                                invalid_df["ASSETTYPE"].value_counts().to_dict()
                            )
                            for at_code, at_count in invalid_dict.items():
                                at_name = "<Invalid Value>"
                                assettype_list.append((at_code, at_name, at_count))
                                # Sort on asset type code
                                sorted_assettype_list = sorted(
                                    assettype_list, key=lambda x: x[0]
                                )
                    subtype_list.append(
                        (
                            subtype_code,
                            subtype_name,
                            subtype_count,
                            sorted_assettype_list,
                        )
                    )

            # Sort on subtype code
            sorted_subtype_list = sorted(subtype_list, key=lambda x: x[0])

        # Add details to data list
        val_tuple = (fds, ds, shape_type, record_count, sorted_subtype_list)
        records.append(val_tuple)
    records.append("")

# Write results to excel
if records:
    row = 2
    for val in records:
        if val == "":
            row += 1
        else:
            for i in range(0, len(val) - 1):
                ws.cell(row=row, column=i + 1, value=val[i])
            row += 1
            if val[-1]:
                for subtype in val[-1]:
                    if subtype[0] != 99999999:
                        ws.cell(row=row, column=5, value=subtype[0])
                    else:
                        # Value 99999999 was used to sort by subtype code
                        # because the value had to be an int for sorting
                        # but want the code to actually show as Null
                        ws.cell(row=row, column=5, value="Null")
                    ws.cell(row=row, column=6, value=subtype[1])
                    ws.cell(row=row, column=7, value=subtype[2])
                    row += 1
                    if include_assettypes:
                        assettype_list = subtype[3]
                        for at_values in assettype_list:
                            ws.cell(row=row, column=8, value=at_values[0])
                            ws.cell(row=row, column=9, value=at_values[1])
                            ws.cell(row=row, column=10, value=at_values[2])
                            row += 1

    # Update formatting for record count columns
    for cell in ws["D"]:
        cell.number_format = "#,##0"

    for cell in ws["G"]:
        cell.number_format = "#,##0"

    if include_assettypes:
        for cell in ws["J"]:
            cell.number_format = "#,##0"

    # Center code fields
    for row in ws.iter_rows(min_row=1, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=1, min_col=8, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Bold and freeze first row
    bold_font = openpyxl.styles.Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    ws.freeze_panes = "A2"

    # Apply autofit to all columns
    autofit_column_widths(ws)

    # Save excel
    wb.save(out_xls)

    # Start file
    os.startfile(out_xls)
