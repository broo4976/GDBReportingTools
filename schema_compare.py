"""
schema_compare.py
Brooke Reams - breams@esri.com
Aug. 5, 2024
Compares the schema of two geodatabases and outputs the differences to an Excel file.

Updates:
9/12/2025:      Fixed case sensitive field names in compare (ignore case) and switched
                base/test columns for fields.
9/17/2025:      Cleaned up unnecessary for loops in fields and removed any duplicate field
                names (duplicates coming from index section of xml)
9/26/2025:      Added configurable logic to only compare field length if field type is string
                (can toggle option on/off with IGNORE_LEN_NON_TEXT_FIELDS bool where True ignores
                comparison for non-string fields and False performs comparison).
                Fixed wording in subtype compare (was printing 'Domain' rather than 'Subtype'
                in output Excel file).
                Swapped base/test for subtype property in fields.
12/11/2025:     Added HasZ and HasM as properties to compare on feature classes.
12/11/2025:     Added parameter to allow user to skip comparison between some properties.
12/12/2025:     Added item (feature class, table, domain, etc) to for missing/additional to
                Base/Test column in output spreadsheet.

"""

from lxml import etree
import operator
import os
import openpyxl
from openpyxl.styles import Alignment

import arcpy


def log_it(message):
    print(message)
    arcpy.AddMessage(message)


def swap_key_value(d):
    d_swap = {}
    for key, val in d.items():
        d_swap[val] = key

    return d_swap


def get_domain_properties(tree):
    for node in tree.iter("Domains"):
        domain_list = []
        for elem in node.iter("Domain"):
            domain_dict = {}
            cv_dict = {}
            for domain_prop in list(elem.iter()):
                if domain_prop.tag in [
                    "DomainName",
                    "FieldType",
                    "MergePolicy",
                    "SplitPolicy",
                ]:
                    domain_dict[domain_prop.tag] = domain_prop.text
                elif domain_prop.tag == "CodedValues":
                    coded_values = domain_prop.getchildren()
                    for cv in coded_values:
                        for cv_prop in cv.iter():
                            if cv_prop.tag == "Name":
                                name = cv_prop.text
                            elif cv_prop.tag == "Code":
                                cv_dict[name] = cv_prop.text
                    sorted_cv_dict = dict(sorted(cv_dict.items()))
                    domain_dict["CodedValues"] = sorted_cv_dict
                elif domain_prop.tag == "MaxValue":
                    max_value = domain_prop.text
                elif domain_prop.tag == "MinValue":
                    min_value = domain_prop.text
                    domain_dict["Range"] = "{} - {}".format(min_value, max_value)
            domain_list.append(domain_dict)

    # Sort domain list alphabetically by domain name
    domain_list.sort(key=operator.itemgetter("DomainName"))

    return domain_list


def compare_domains(tree_base, tree_test):
    domain_list_base = get_domain_properties(tree_base)
    domain_list_test = get_domain_properties(tree_test)
    if domain_list_base == domain_list_test:
        return ([], {}, [])
    else:
        # Find domains in base that aren't identical to test
        domain_diff_test = [i for i in domain_list_test if i not in domain_list_base]
        # Find missing domains
        domain_names_test = [i["DomainName"] for i in domain_list_test]
        domain_names_base = [i["DomainName"] for i in domain_list_base]
        domain_miss = [
            i for i in domain_list_base if i["DomainName"] not in domain_names_test
        ]
        domain_add = [
            i for i in domain_list_test if i["DomainName"] not in domain_names_base
        ]
        # Find mismatch domains
        for domain in domain_add:
            domain_diff_test.remove(domain)
        domain_diff = {}
        for domain in domain_diff_test:
            domain_name = domain["DomainName"]
            # Get domain out of base list
            domain_base = [
                d for d in domain_list_base if d["DomainName"] == domain_name
            ][0]
            for key, val in domain.items():
                # Get mismatch in property values
                if key != "CodedValues":
                    if val != domain_base[key]:
                        domain_diff[domain_name] = [
                            (
                                "Domain has mismatch {} property".format(key),
                                domain_base[key],
                                val,
                            )
                        ]
                # Get mismatched coded values
                else:
                    if val != domain_base[key]:
                        base_diff = (
                            dict(set(domain_base[key].items()) - set(val.items()))
                            if bool(
                                dict(set(domain_base[key].items()) - set(val.items()))
                            )
                            else ""
                        )
                        test_diff = (
                            dict(set(val.items()) - set(domain_base[key].items()))
                            if bool(
                                dict(set(val.items()) - set(domain_base[key].items()))
                            )
                            else ""
                        )
                        if base_diff != "" and test_diff == "":
                            adj = "additional and missing"
                            base_diff = swap_key_value(base_diff)
                            test_diff = swap_key_value(test_diff)
                        elif base_diff != "":
                            base_diff = swap_key_value(base_diff)
                            adj = "missing"
                        elif test_diff != "":
                            test_diff = swap_key_value(test_diff)
                            adj = "additional"

                        if domain_name in domain_diff.keys():
                            domain_diff[domain_name].append(
                                (
                                    "Domain has {} CodedValues".format(adj),
                                    str(base_diff),
                                    str(test_diff),
                                )
                            )
                        else:
                            domain_diff[domain_name] = [
                                (
                                    "Domain has {} CodedValues".format(adj),
                                    str(base_diff),
                                    str(test_diff),
                                )
                            ]
        return (domain_miss, domain_diff, domain_add)


def get_dataset_properties(
    tree,
    ds_type,
    ignore_ds_alias,
    ignore_fld_alias,
    ignore_hasm,
    ignore_hasz,
):
    dataset_list = []
    for elem in tree.iter("DataElement"):
        if elem.attrib.values()[0] == ds_type:
            dataset_dict = {"SubtypeFieldName": ""}
            subtype_name_fld = ""
            subtype_default_code = ""
            ##dataset_dict["SubtypeFieldName"] = ""
            flds_list = []
            subtype_list = []
            for ds_prop in list(elem.iter()):
                flds_dict = {}
                if (
                    ds_prop.tag
                    in [
                        "Name",
                        "Versioned",
                        "CanVersion",
                        "ConfigurationKeyword",
                        "ShapeType",
                    ]
                    and ds_prop.getparent().tag == "DataElement"
                ):
                    dataset_dict[ds_prop.tag] = ds_prop.text
                elif (
                    ds_prop.tag == "AliasName"
                    and not ignore_ds_alias
                    and ds_prop.getparent().tag == "DataElement"
                ):
                    dataset_dict[ds_prop.tag] = ds_prop.text
                elif (
                    ds_prop.tag == "HasM"
                    and not ignore_hasm
                    and ds_prop.getparent().tag == "DataElement"
                ):
                    dataset_dict[ds_prop.tag] = ds_prop.text
                elif (
                    ds_prop.tag == "HasZ"
                    and not ignore_hasz
                    and ds_prop.getparent().tag == "DataElement"
                ):
                    dataset_dict[ds_prop.tag] = ds_prop.text
                elif ds_prop.tag == "Subtype":
                    subtype_dict = {}
                    for s in ds_prop.iter():
                        if s.tag in ["SubtypeName", "SubtypeCode"]:
                            subtype_dict[s.tag] = s.text
                    subtype_list.append(subtype_dict)
                elif ds_prop.tag == "SubtypeFieldName":
                    dataset_dict[ds_prop.tag] = ds_prop.text
                    subtype_name_fld = ds_prop.text
                elif ds_prop.tag == "DefaultSubtypeCode":
                    subtype_default_code = ds_prop.text
                    ##dataset_dict[ds_prop.tag] = ds_prop.text
                elif ds_prop.tag == "Field":
                    flds_data = ds_prop.getchildren()
                    for fld_prop in flds_data:
                        flds_dict["Domain"] = ""
                        if fld_prop.tag in [
                            "Type",
                            "IsNullable",
                            "Length",
                            "Precision",
                            "Scale",
                            "Required",
                            "Editable",
                            "DefaultValue",
                        ]:
                            flds_dict[fld_prop.tag] = fld_prop.text
                        elif fld_prop.tag == "Name":
                            flds_dict[fld_prop.tag] = fld_prop.text.lower()
                        elif fld_prop.tag == "AliasName" and not ignore_fld_alias:
                            flds_dict[fld_prop.tag] = fld_prop.text
                        elif fld_prop.tag == "Domain":
                            domain_data = fld_prop.getchildren()
                            for domain_prop in domain_data:
                                if domain_prop.tag == "DomainName":
                                    flds_dict["Domain"] = domain_prop.text
                    if (
                        flds_dict["Type"] != "esriFieldTypeString"
                        and IGNORE_LEN_NON_TEXT_FIELDS == True
                    ):
                        flds_dict.__delitem__("Length")
                    if flds_dict not in flds_list:
                        flds_list.append(flds_dict)
                    dataset_dict["Fields"] = flds_list
            dataset_dict["Subtypes"] = subtype_list
            dataset_dict["SubtypeInfo"] = {subtype_name_fld: subtype_default_code}
            dataset_list.append(dataset_dict)
            dataset_list.sort(key=operator.itemgetter("Name"))

    return dataset_list


def compare_datasets(
    tree_base,
    tree_test,
    ds_type,
    name,
    ignore_ds_alias,
    ignore_fld_alias,
    ignore_hasm=True,
    ignore_hasz=True,
):
    ds_list_base = get_dataset_properties(
        tree_base,
        ds_type,
        ignore_ds_alias,
        ignore_fld_alias,
        ignore_hasm,
        ignore_hasz,
    )
    ds_list_test = get_dataset_properties(
        tree_test,
        ds_type,
        ignore_ds_alias,
        ignore_fld_alias,
        ignore_hasm,
        ignore_hasz,
    )
    if ds_list_base == ds_list_test:
        return ([], {}, [])
    else:
        # Find datasets in base that aren't identical to test
        ds_diff_test = [i for i in ds_list_test if i not in ds_list_base]
        # Find missing datasets
        ds_names_test = [i["Name"] for i in ds_list_test]
        ds_names_base = [i["Name"] for i in ds_list_base]
        ds_miss = [i for i in ds_list_base if i["Name"] not in ds_names_test]
        ds_add = [i for i in ds_list_test if i["Name"] not in ds_names_base]
        # Find mismatch datasets
        for ds in ds_add:
            ds_diff_test.remove(ds)
        ds_diff = {}
        for ds in ds_diff_test:
            ds_name = ds["Name"]
            # Get dataset out of base list
            ds_base = [d for d in ds_list_base if d["Name"] == ds_name][0]
            for key, val in ds.items():
                ##log_it(val)
                if key == "SubtypeInfo":
                    subtype_info_base = ds_base["SubtypeInfo"]
                    if list(subtype_info_base.keys())[0] == list(val.keys())[0]:
                        if list(subtype_info_base.values())[0] != list(val.values())[0]:
                            if ds_name in ds_diff.keys():
                                ds_diff[ds_name].append(
                                    (
                                        "{} has mismatch DefaultSubtypeCode property".format(
                                            name, key
                                        ),
                                        list(subtype_info_base.keys())[0],
                                        list(val.keys())[0],
                                    )
                                )
                            else:
                                ds_diff[ds_name] = (
                                    "{} has mismatch DefaultSubtypeCode property".format(
                                        name, key
                                    ),
                                    list(subtype_info_base.keys())[0],
                                    list(val.keys())[0],
                                )
                elif key == "Subtypes":
                    st_base_list = ds_base["Subtypes"]
                    st_names_test = [st["SubtypeName"] for st in val]
                    st_names_base = [st["SubtypeName"] for st in st_base_list]
                    st_miss = [
                        i for i in st_base_list if i["SubtypeName"] not in st_names_test
                    ]
                    st_add = [i for i in val if i["SubtypeName"] not in st_names_base]
                    for st in st_add:
                        val.remove(st)
                        if ds_name in ds_diff.keys():
                            ds_diff[ds_name].append(
                                ("Additional subtype", "", st["SubtypeName"])
                            )
                        else:
                            ds_diff[ds_name] = (
                                "Additional subtype",
                                "",
                                st["SubtypeName"],
                            )

                    for st in st_miss:
                        if ds_name in ds_diff.keys():
                            ds_diff[ds_name].append(
                                ("Missing subtype", st["SubtypeName"], "")
                            )
                        else:
                            ds_diff[ds_name] = [
                                ("Missing subtype", st["SubtypeName"], "")
                            ]

                elif key == "Fields":
                    # Get base fields
                    flds_base_list = ds_base["Fields"]
                    flds_names_test = [fld["Name"] for fld in val]
                    flds_names_base = [fld["Name"] for fld in flds_base_list]
                    flds_miss = [
                        i for i in flds_base_list if i["Name"] not in flds_names_test
                    ]
                    flds_add = [i for i in val if i["Name"] not in flds_names_base]
                    for fld in flds_add:
                        val.remove(fld)
                        if ds_name in ds_diff.keys():
                            ds_diff[ds_name].append(
                                ("Additional field", "", fld["Name"])
                            )
                        else:
                            ds_diff[ds_name] = [("Additional field", "", fld["Name"])]

                    for fld in flds_miss:
                        if ds_name in ds_diff.keys():
                            ds_diff[ds_name].append(("Missing field", fld["Name"], ""))
                        else:
                            ds_diff[ds_name] = [("Missing field", fld["Name"], "")]

                    for fld_prop in val:
                        fld_name = fld_prop["Name"]
                        for flds_base in flds_base_list:
                            if flds_base["Name"] == fld_name:
                                # Check if properties of fields are the same
                                for base_key, base_val in flds_base.items():
                                    ##log_it("{}: {}".format(base_key, base_val))
                                    for test_key, test_val in fld_prop.items():
                                        if base_key == test_key:
                                            if base_val != test_val:
                                                if ds_name in ds_diff.keys():
                                                    ds_diff[ds_name].append(
                                                        (
                                                            "{} field has different values for {}".format(
                                                                fld_name, base_key
                                                            ),
                                                            str(base_val),
                                                            str(test_val),
                                                        )
                                                    )
                                                else:
                                                    ds_diff[ds_name] = [
                                                        (
                                                            "{} field has different values for {}".format(
                                                                fld_name, base_key
                                                            ),
                                                            str(base_val),
                                                            str(test_val),
                                                        )
                                                    ]
                else:
                    # Get mismatch in property values
                    if val != ds_base[key]:
                        if ds_name in ds_diff.keys():
                            ds_diff[ds_name].append(
                                (
                                    "{} has mismatch {} property".format(name, key),
                                    ds_base[key],
                                    val,
                                )
                            )
                        else:
                            ds_diff[ds_name] = [
                                (
                                    "{} has mismatch {} property".format(name, key),
                                    ds_base[key],
                                    val,
                                )
                            ]

        return (ds_miss, ds_diff, ds_add)


def get_rc_properties(tree, ds_type):
    rc_list = []
    for elem in tree.iter("DataElement"):
        if elem.attrib.values()[0] == ds_type:
            rc_dict = {}
            for rc_prop in list(elem.getchildren()):
                if rc_prop.tag in [
                    "Name",
                    "Versioned",
                    "CanVersion",
                    "ConfigurationKeyword",
                    "HasOID",
                    "OIDFieldName",
                    "Fields",
                    "Cardinality",
                    "IsComposite",
                    "OriginClassNames",
                    "DestinationClassNames",
                    "KeyType",
                    "ClassKey",
                    "IsReflexive",
                    "OriginClassKeys",
                    "RelationshipRules",
                    "IsAttachmentRelationship",
                ]:
                    rc_dict[rc_prop.tag] = rc_prop.text
            rc_list.append(rc_dict)
            rc_list.sort(key=operator.itemgetter("Name"))

    return rc_list


def compare_relationship_classes(tree_base, tree_test, ds_type, name):
    rc_list_base = get_rc_properties(tree_base, ds_type)
    rc_list_test = get_rc_properties(tree_test, ds_type)
    if rc_list_base == rc_list_test:
        return ([], {}, [])
    else:
        # Find relationship classes in base that aren't identical to test
        rc_diff_test = [i for i in rc_list_test if i not in rc_list_base]
        # Find missing relationship classes
        rc_names_test = [i["Name"] for i in rc_list_test]
        rc_names_base = [i["Name"] for i in rc_list_base]
        rc_miss = [i for i in rc_list_base if i["Name"] not in rc_names_test]
        rc_add = [i for i in rc_list_test if i["Name"] not in rc_names_base]
        # Find mismatch relationship classes
        for rc in rc_add:
            rc_diff_test.remove(rc)
        rc_diff = {}
        for rc in rc_diff_test:
            rc_name = rc["Name"]
            # Get relationship class out of base list
            rc_base = [r for r in rc_list_base if r["Name"] == rc_name][0]
            for key, val in rc.items():
                # Get mismatch in property values
                if val != rc_base[key]:
                    rc_diff[rc_name] = [
                        (
                            "{} has mismatch {} property".format(name, key),
                            rc_base[key],
                            val,
                        )
                    ]

    return (rc_miss, rc_diff, rc_add)


def get_fds_properties(tree, ds_type):
    fds_list = []
    for elem in tree.iter("DataElement"):
        if elem.attrib.values()[0] == ds_type:
            fds_dict = {"Children": []}
            for fds_prop in list(elem.iter()):
                if "esriDTFeatureDataset" in [
                    i.text for i in fds_prop.getchildren() if i.tag == "DatasetType"
                ]:
                    children = fds_prop.getchildren()
                    for child in children:
                        if child.tag in [
                            "DatasetType",
                            "Name",
                            "Versioned",
                            "CanVersion",
                            "Extent",
                            "SpatialReference",
                        ]:
                            fds_dict[child.tag] = child.text
                        elif child.tag == "Children":
                            grandchildren = child.getchildren()
                            for grandchild in grandchildren:
                                if (
                                    grandchild.tag == "DataElement"
                                    and grandchild.attrib.values()[0]
                                    == "esri:DEFeatureClass"
                                ):
                                    greatgrandchildren = grandchild.getchildren()
                                    for greatgrandchild in greatgrandchildren:
                                        if greatgrandchild.tag == "Name":
                                            fds_dict["Children"].append(
                                                greatgrandchild.text
                                            )
                                    fds_dict["Children"].sort()
                    fds_list.append(fds_dict)
                    fds_list.sort(key=operator.itemgetter("Name"))

    return fds_list


def compare_fds(tree_base, tree_test, ds_type, name):
    fds_list_base = get_fds_properties(tree_base, ds_type)
    fds_list_test = get_fds_properties(tree_test, ds_type)
    if fds_list_base == fds_list_test:
        return ([], {}, [])
    else:
        # Find feature datasets in base that aren't identical to test
        fds_diff_test = [i for i in fds_list_test if i not in fds_list_base]
        # Find missing feature datasets
        fds_names_test = [i["Name"] for i in fds_list_test]
        fds_names_base = [i["Name"] for i in fds_list_base]
        fds_miss = [i for i in fds_list_base if i["Name"] not in fds_names_test]
        fds_add = [i for i in fds_list_test if i["Name"] not in fds_names_base]
        # Find mismatch feature datasets
        for fds in fds_add:
            fds_diff_test.remove(fds)
        fds_diff = {}
        for fds in fds_diff_test:
            fds_name = fds["Name"]
            # Get feature datasets out of base list
            fds_base = [f for f in fds_list_base if f["Name"] == fds_name][0]
            for key, val in fds.items():
                if key != "Children":
                    # Get mismatch in property values
                    if val != fds_base[key]:
                        if fds_name not in fds_diff.keys():
                            fds_diff[fds_name] = [
                                (
                                    "{} has mismatch {} property".format(name, key),
                                    fds_base[key],
                                    val,
                                )
                            ]
                        else:
                            fds_diff[fds_name].append(
                                (
                                    "{} has mismatch {} property".format(name, key),
                                    fds_base[key],
                                    val,
                                )
                            )
    return (fds_miss, fds_diff, fds_add)


def get_topo_properties(tree, ds_type):
    # Get dictionary of feature class ids: names
    fc_dict = {}
    for elem in tree.iter("DataElement"):
        if elem.attrib.values()[0] == "esri:DEFeatureClass":
            for prop in list(elem.getchildren()):
                if prop.tag == "Name":
                    name = prop.text
                elif prop.tag == "DSID":
                    fc_dict[prop.text] = name

    # Get topo properties
    topo_list = []
    for elem in tree.iter("DataElement"):
        if elem.attrib.values()[0] == ds_type:
            topo_dict = {}
            fc_list = []
            rule_list = []
            for topo_prop in list(elem.getchildren()):
                if topo_prop.tag in [
                    "Name",
                    "ClusterTolerance",
                    "ZClusterTolerance",
                    "MaxGeneratedErrorCount",
                ]:
                    topo_dict[topo_prop.tag] = topo_prop.text
                elif topo_prop.tag == "FeatureClassNames":
                    for fc in topo_prop.iter():
                        if fc.tag == "Name":
                            fc_list.append(fc.text)
                        fc_list.sort()
                elif topo_prop.tag == "TopologyRules":
                    for topo_rules in list(topo_prop.getchildren()):
                        rule_dict = {}
                        for topo_rule in list(topo_rules.getchildren()):
                            if topo_rule.tag == "TopologyRuleType":
                                rule_dict[topo_rule.tag] = topo_rule.text
                            elif topo_rule.tag in [
                                "OriginClassID",
                                "DestinationClassID",
                            ]:
                                rule_dict[topo_rule.tag] = fc_dict[topo_rule.text]
                            elif topo_rule.tag in [
                                "OriginSubtype",
                                "DestinationSubtype",
                            ]:
                                rule_dict[topo_rule.tag] = topo_rule.text
                        rule_list.append(rule_dict)

            topo_dict["FeatureClassNames"] = fc_list
            topo_dict["TopologyRules"] = sorted(
                rule_list,
                key=lambda x: (
                    x["OriginClassID"],
                    x["TopologyRuleType"],
                    x["DestinationClassID"],
                    x["OriginSubtype"],
                    x["DestinationSubtype"],
                ),
            )
            topo_list.append(topo_dict)
            topo_list.sort(key=operator.itemgetter("Name"))

    return topo_list


def compare_topo(tree_base, tree_test, ds_type, name):
    topo_list_base = get_topo_properties(tree_base, ds_type)
    topo_list_test = get_topo_properties(tree_test, ds_type)

    if topo_list_base == topo_list_test:
        return ([], {}, [])
    else:
        # Find topo datasets in base that aren't identical to test
        topo_diff_test = [i for i in topo_list_test if i not in topo_list_base]
        # Find missing topo datasets
        topo_names_test = [i["Name"] for i in topo_list_test]
        topo_names_base = [i["Name"] for i in topo_list_base]
        topo_miss = [i for i in topo_list_base if i["Name"] not in topo_names_test]
        topo_add = [i for i in topo_list_test if i["Name"] not in topo_names_base]
        # Find mismatch topo datasets
        for topo in topo_add:
            topo_diff_test.remove(topo)
        topo_diff = {}
        for topo in topo_diff_test:
            topo_name = topo["Name"]
            # Get topo datasets out of base list
            topo_base = [t for t in topo_list_base if t["Name"] == topo_name][0]
            for key, val in topo.items():
                if key not in ["FeatureClassNames", "TopologyRules"]:
                    # Get mismatch in property values
                    if val != topo_base[key]:
                        if topo_name not in topo_diff.keys():
                            topo_diff[topo_name] = [
                                (
                                    "{} has mismatch {} property".format(name, key),
                                    topo_base[key],
                                    val,
                                )
                            ]
                        else:
                            topo_diff[topo_name].append(
                                (
                                    "{} has mismatch {} property".format(name, key),
                                    topo_base[key],
                                    val,
                                )
                            )
                elif key == "FeatureClassNames":
                    if val != topo_base[key]:
                        miss_fcs = list(set(topo_base[key]) - set(val))
                        add_fcs = list(set(val) - set(topo_base[key]))
                        if len(miss_fcs) > 0:
                            for fc in miss_fcs:
                                if topo_name not in topo_diff.keys():
                                    topo_diff[topo_name] = [
                                        ("Missing feature class in topology", fc, None)
                                    ]
                                else:
                                    topo_diff[topo_name].append(
                                        ("Missing feature class in topology", fc, None)
                                    )

                        if len(add_fcs) > 0:
                            for fc in add_fcs:
                                if topo_name not in topo_diff.keys():
                                    topo_diff[topo_name] = [
                                        (
                                            "Additional feature class in topology",
                                            None,
                                            fc,
                                        )
                                    ]
                                else:
                                    topo_diff[topo_name].append(
                                        (
                                            "Additional feature class in topology",
                                            None,
                                            fc,
                                        )
                                    )
                elif key == "TopologyRules":
                    if val != topo_base[key]:
                        miss_rules = []
                        add_rules = []
                        for rule in topo_base[key]:
                            if rule not in val:
                                miss_rules.append(rule)

                        for rule in val:
                            if rule not in topo_base[key]:
                                add_rules.append(rule)

                        if len(miss_rules) > 0:
                            for rule in miss_rules:
                                if topo_name not in topo_diff.keys():
                                    topo_diff[topo_name] = [
                                        (
                                            "Missing topology rule",
                                            str(rule).replace("'", ""),
                                            None,
                                        )
                                    ]
                                else:
                                    topo_diff[topo_name].append(
                                        (
                                            "Missing topology rule",
                                            str(rule).replace("'", ""),
                                            None,
                                        )
                                    )
                        if len(add_rules) > 0:
                            for rule in add_rules:
                                if topo_name not in topo_diff.keys():
                                    topo_diff[topo_name] = [
                                        (
                                            "Additional topology rule",
                                            None,
                                            str(rule).replace("'", ""),
                                        )
                                    ]
                                else:
                                    topo_diff[topo_name].append(
                                        (
                                            "Additional topology rule",
                                            None,
                                            str(rule).replace("'", ""),
                                        )
                                    )

    return (topo_miss, topo_diff, topo_add)


def get_attr_rules_properties(tree, ds_type):
    # Get dictionary of attribute rule names
    attr_rules_dict = {}

    for elem in tree.iter("DataElement"):
        if elem.attrib.values()[0] in ds_type:
            for prop in list(elem.getchildren()):
                if prop.tag == "Name":
                    ds_name = prop.text
                if prop.tag == "AttributeRules":
                    for a in list(prop.getchildren()):
                        if a.tag == "AttributeRule":
                            r_dict = {}
                            r_dict["DatasetName"] = ds_name
                            for i in list(a.getchildren()):
                                if i.tag == "Name":
                                    key = f"{ds_name}: {i.text}"
                                elif i.tag in [
                                    "Type",
                                    "FieldName",
                                    "SubtypeCode",
                                    "Description",
                                    "UserEditable",
                                    "IsEnabled",
                                    "ReferencesExternalService",
                                    "ExcludeFromClientEvaluation",
                                    "ScriptExpression",
                                    "TriggeringEvents",
                                ]:
                                    r_dict[i.tag] = i.text
                            sorted_dict = dict(sorted(r_dict.items()))
                        attr_rules_dict[key] = sorted_dict

    keys_sorted = sorted(attr_rules_dict.keys())
    attr_rules_dict_sorted = {key: attr_rules_dict[key] for key in keys_sorted}
    return attr_rules_dict_sorted


def compare_attr_rules(tree_base, tree_test, ds_type, name):
    attr_rules_list_base = get_attr_rules_properties(tree_base, ds_type)
    attr_rules_list_test = get_attr_rules_properties(tree_test, ds_type)

    if attr_rules_list_base == attr_rules_list_test:
        return ([], {}, [])
    else:
        # Check for keys (attribute rule name/fc) missing in base
        base_keys = list(attr_rules_list_base.keys())
        base_keys.sort()
        test_keys = list(attr_rules_list_test.keys())
        test_keys.sort()

        if base_keys != test_keys:
            # Check for attribute rules missing and additional
            attr_miss = list(set(base_keys) - set(test_keys))
            attr_add = list(set(test_keys) - set(base_keys))

        # Check for differences in attribute rules
        attr_diff = {}
        attr_diff_items = [i for i in base_keys if i in test_keys]
        for i in attr_diff_items:
            prop_base = attr_rules_list_base[i]
            prop_test = attr_rules_list_test[i]

            if prop_base != prop_test:
                for key, val in prop_base.items():
                    if val != prop_test[key]:
                        if i not in attr_diff.keys():
                            domain_name_list = i.split(":")
                            domain_name = "".join(domain_name_list[1:]).strip()
                            attr_diff[i] = [
                                (
                                    "{} has mismatch {} property".format(
                                        domain_name, key
                                    ),
                                    prop_base[key],
                                    prop_test[key],
                                )
                            ]
                        else:
                            attr_diff[i].append(
                                (
                                    "{} has mismatch {} property".format(
                                        domain_name, key
                                    ),
                                    prop_base[key],
                                    prop_test[key],
                                )
                            )
    return (attr_miss, attr_diff, attr_add)


def write_results_to_xls(
    wb, sheet_name, item_type, dict_key, miss_list, diff_dict, adds_list
):
    # Write results to excel
    ws = wb.create_sheet(sheet_name)
    bold_font = openpyxl.styles.Font(bold=True)
    for cell in ws["A1:D1"][0]:
        cell.font = bold_font
    ws["A1"] = "{} Name".format(item_type)
    ws["B1"] = "Difference"
    ws["C1"] = "Base"
    ws["D1"] = "Test"

    row = 2
    for item in miss_list:
        if dict_key:
            ws.cell(row=row, column=1).value = item[dict_key]
        else:
            ws.cell(row=row, column=1).value = item
        ws.cell(row=row, column=2).value = "Missing {}".format(item_type)
        ws.cell(row=row, column=3).value = ws.cell(row=row, column=1).value
        row += 1

    for key, val_list in diff_dict.items():
        ws.cell(row=row, column=1).value = key
        start_row = row
        for val in val_list:
            col = 2
            for item in val:
                ws.cell(row=row, column=col).value = item
                col += 1
            row += 1
        # If more than one issue for an item, merge cells
        if start_row + 1 < row:
            ws.merge_cells(
                start_row=start_row, start_column=1, end_row=row - 1, end_column=1
            )
            # Vertically align text merged cell
            merged_cell = ws.cell(row=start_row, column=1)
            merged_cell.alignment = Alignment(vertical="center")

    for item in adds_list:
        if dict_key:
            ws.cell(row=row, column=1).value = item[dict_key]
        else:
            ws.cell(row=row, column=1).value = item
        ws.cell(row=row, column=2).value = "Additional {}".format(item_type)
        ws.cell(row=row, column=4).value = ws.cell(row=row, column=1).value
        row += 1


# Configuration
IGNORE_LEN_NON_TEXT_FIELDS = (
    True  # Flag to skip the comparison of length property of non-text fields
)

# Get xml files
xml_file_base = arcpy.GetParameterAsText(0)
xml_file_test = arcpy.GetParameterAsText(1)

# Output xls file
out_xls = arcpy.GetParameterAsText(2)

# Optional properties to ignore
ignore_str = arcpy.GetParameterAsText(3)

# Create dictionary to store ignore values and bools
ignore_dict = {
    "Feature Class/Table Alias": False,
    "Field Alias": False,
    "Has M": False,
    "Has Z": False,
    "Domains": False,
    "Topology": False,
}

# Get value table values
if ignore_str:
    ignore_list = ignore_str.split(";")
    for i in ignore_list:
        i = i.replace("'", "")
        if i in ignore_dict.keys():
            ignore_dict[i] = True
        else:
            arcpy.AddWarning(f"{i} Ignore Property unknown")

ignore_ds_alias = ignore_dict["Feature Class/Table Alias"]
ignore_fld_alias = ignore_dict["Field Alias"]
ignore_hasm = ignore_dict["Has M"]
ignore_hasz = ignore_dict["Has Z"]
ignore_domains = ignore_dict["Domains"]
ignore_topology = ignore_dict["Topology"]

# Parse xml files
tree_base = etree.parse(xml_file_base)
tree_test = etree.parse(xml_file_test)


# Open new excel spreadsheet
wb = openpyxl.Workbook()
wb.remove(wb["Sheet"])
save_wb = False


# COMPARE FEATURE DATASETS
log_it("Comparing feature datasets")
fds_miss, fds_diff, fds_add = compare_fds(
    tree_base, tree_test, "esri:DEFeatureDataset", "Feature Dataset"
)
if fds_miss or fds_diff or fds_add:
    save_wb = True
    write_results_to_xls(
        wb, "Feature Datasets", "Feature Dataset", "Name", fds_miss, fds_diff, fds_add
    )


# COMPARE FEATURE CLASSES
log_it("Comparing feature classes")
fc_miss, fc_diff, fc_add = compare_datasets(
    tree_base,
    tree_test,
    "esri:DEFeatureClass",
    "Feature Class",
    ignore_ds_alias,
    ignore_fld_alias,
    ignore_hasm,
    ignore_hasz,
)
if fc_miss or fc_diff or fc_add:
    save_wb = True
    write_results_to_xls(
        wb, "Feature Classes", "Feature Class", "Name", fc_miss, fc_diff, fc_add
    )


# COMPARE TABLES
log_it("Comparing tables")
tbl_miss, tbl_diff, tbl_add = compare_datasets(
    tree_base,
    tree_test,
    "esri:DETable",
    "Table",
    ignore_ds_alias,
    ignore_fld_alias,
    ignore_domains,
)
if tbl_miss or tbl_diff or tbl_add:
    save_wb = True
    write_results_to_xls(wb, "Tables", "Table", "Name", tbl_miss, tbl_diff, tbl_add)


# COMPARE RELATIONSHIP CLASSES
log_it("Comparing relationship classes")
rc_miss, rc_diff, rc_add = compare_relationship_classes(
    tree_base, tree_test, "esri:DERelationshipClass", "Relationship Class"
)
if rc_miss or rc_diff or rc_add:
    save_wb = True
    write_results_to_xls(
        wb,
        "Relationship Classes",
        "Relationship Class",
        "Name",
        rc_miss,
        rc_diff,
        rc_add,
    )


if not ignore_domains:
    # COMPARE DOMAINS
    log_it("Comparing domains")
    domain_miss, domain_diff, domain_add = compare_domains(tree_base, tree_test)
    if domain_miss or domain_diff or domain_add:
        save_wb = True
        write_results_to_xls(
            wb, "Domains", "Domain", "DomainName", domain_miss, domain_diff, domain_add
        )


if not ignore_topology:
    # COMPARE TOPOLOGIES
    log_it("Comparing topologies")
    topo_miss, topo_diff, topo_add = compare_topo(
        tree_base, tree_test, "esri:DETopology", "Topology"
    )
    if topo_miss or topo_diff or topo_add:
        save_wb = True
        write_results_to_xls(
            wb, "Topologies", "Topology", "Name", topo_miss, topo_diff, topo_add
        )


# COMPARE ATTRIBUTE RULES
log_it("Comparing attribute rules")
attr_miss, attr_diff, attr_add = compare_attr_rules(
    tree_base, tree_test, ["esri:DEFeatureClass", "esri:DETable"], "Attribute Rules"
)
if attr_miss or attr_diff or attr_add:
    save_wb = True
    write_results_to_xls(
        wb, "Attribute Rules", "Attribute Rule", "", attr_miss, attr_diff, attr_add
    )


if save_wb:
    log_it("Differences found.  Creating Diff xlsx.")
    # Save excel file
    wb.save(out_xls)
    # Open excel file
    os.startfile(out_xls)
else:
    log_it("There were no differences found between schemas.")
