"""
Brooke Reams - breams@esri.com
Oct. 16, 2025

Description:
Loops through all feature classes and tables at the root
and feature dataset level and reports globlaid, editor
tracking, and coordinate system properties on each dataset.

ArcGIS Pro 3.5.2
Python 3.11.11

Updates:

"""

import arcpy
import os
import openpyxl
from openpyxl.utils import get_column_letter

# Overwrite existing output
arcpy.env.overwriteOutput = 1

def log_it(message):
    print(message)
    arcpy.AddMessage(message)


def autofit_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column) # Get column letter from the first cell in the column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError: # Handle cases where cell.value might be None or not easily convertible to string
                pass
        adjusted_width = (max_length + 2) # Add some padding
        ws.column_dimensions[column].width = adjusted_width


# Tool inputs
in_ws = arcpy.GetParameterAsText(0)
out_xls = arcpy.GetParameterAsText(1)

# Create new workbook and define header
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "Feature Dataset"
ws["B1"] = "Feature Class/Table"
ws["C1"] = "Has Esri GlobalID"
ws["D1"] = "GlobalID Field Name"
ws["E1"] = "GlobalID Type"
ws["F1"] = "Editor Tracking Enabled"
ws["G1"] = "Creator Field"
ws["H1"] = "Date Created Field"
ws["I1"] = "Edited By Field"
ws["J1"] = "Edited Date Field"
ws["K1"] = "Coordinate System"
ws["L1"] = "WKID"
ws["M1"] = "Linear Units"
ws["N1"] = "Vertical CS"

# Set workspace environment
arcpy.env.workspace = in_ws

# Initialize list to store data properties
records = []

# Loop through all feature classes/tables in gdb and in all feature datasets
fds_list = arcpy.ListDatasets(feature_type="Feature")
fds_list.sort()
fds_list.append("")
for fds in fds_list:
    # Get feature classes/tables
    if fds == "":
        log_it(f"Processing stand-alone datasets")
        fds = "<standalone>"
        ds_list = arcpy.ListFeatureClasses() + arcpy.ListTables()
    else:
        log_it(f"Processing feature dataset: {fds}") 
        ds_list = arcpy.ListFeatureClasses(feature_dataset=fds)
        
    for ds in ds_list:
        log_it(f"Processing dataset: {ds}")
        # Describe dataset to get properties
        desc = arcpy.Describe(ds)
        
        # Get spatial reference properties
        try:
            spatial_ref = desc.spatialReference
            sr_name = spatial_ref.name
            coord_system = spatial_ref.type
            wkid = spatial_ref.factoryCode
            vertical_cs = spatial_ref.VCS
            if coord_system == "Geographic":
                linear_units = "N/A"
            else:
                linear_units = spatial_ref.linearUnitName
            if not vertical_cs:
                vertical_cs = "None"
        except:
            spatial_ref = ""
            sr_name = ""
            coord_system = ""
            wkid = ""
            vertical_cs = ""
            linear_units = ""
            vertical_cs = ""
            pass
            
        # Get GlobalID field info
        has_globalid = desc.HasGlobalID
        globalid_fld = [fld for fld in arcpy.ListFields(ds) if fld.name.lower() == "globalid"]
        if globalid_fld:
            globalid_type = globalid_fld[0].type
            globalid_fld = globalid_fld[0].name
        else:
            globalid_type = ""
            globalid_fld = ""

        # Get Editor Tracking info
        has_editor_tracking = desc.editorTrackingEnabled
        if has_editor_tracking:
            creator_fld = desc.creatorFieldName
            createddate_fld = desc.createdAtFieldName
            editor_fld = desc.editorFieldName
            editdate_fld = desc.editedAtFieldName
        else:
            creator_fld = ""
            createddate_fld = ""
            editor_fld = ""
            editdate_fld = ""
        
        # Add details to data list
        val_tuple = (fds, ds, str(has_globalid), globalid_fld, globalid_type, str(has_editor_tracking), creator_fld, createddate_fld,
                     editor_fld, editdate_fld, coord_system, wkid, linear_units, vertical_cs)
        records.append(val_tuple)
    records.append("")

# Write results to excel
if records:
    row = 2
    for val in records:
        if val == "":
            row+=1
        else:
            for i in range(0, len(val)):
                ws.cell(row=row, column=i+1, value=val[i])
            row+=1

    # Bold and freeze first row
    bold_font = openpyxl.styles.Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    ws.freeze_panes = "A2"

    # Apply autofit to all columns
    autofit_column_widths(ws)

    # Save excel
    wb.save(out_xls)
    
    # Start file
    os.startfile(out_xls)
