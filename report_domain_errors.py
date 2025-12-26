"""
report_domain_errors.py
Brooke Reams - breams@esri.com
Dec. 22, 2025

Description:
Finds values in domain fields that do not conform to domain coded values/ranges.
Feature classes/tables with domain errors are reported in an output Excel file.

ArcGIS Pro 3.6.0
Python 3.13.7

Updates:


"""

import arcpy
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# Overwrite existing output
arcpy.env.overwriteOutput = 1


def log_it(message):
    print(message)
    arcpy.AddMessage(message)


def autofit_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(
            col[0].column
        )  # Get column letter from the first cell in the column
        # Skip column C because it's word wrapped
        if column != "C":
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (
                    TypeError
                ):  # Handle cases where cell.value might be None or not easily convertible to string
                    pass
            adjusted_width = max_length + 2  # Add some padding
            ws.column_dimensions[column].width = adjusted_width


# Tool inputs
in_ws = arcpy.GetParameterAsText(0)
ds_list = arcpy.GetParameter(1)
out_xls = arcpy.GetParameterAsText(2)

# Get list of all domains in workspace
log_it("Retrieving all domains from workspace")
domain_list = arcpy.da.ListDomains(in_ws)
# Convert domain list into dictionary {domain name: domain object}
domain_dict = {}
for domain in domain_list:
    domain_dict[domain.name] = domain

# Convert dataset list to dictionary {fds: [list of fcs]}
ds_dict = {}
for ds in ds_list:
    if not "/" in ds:
        if "stand_alone" in ds_dict.keys():
            ds_dict["stand_alone"].append(ds)
        else:
            ds_dict["stand_alone"] = [ds]
    else:
        fds = ds.split("/")[0]
        if fds in ds_dict.keys():
            ds_dict[fds].append(ds.split("/")[-1])
        else:
            ds_dict[fds] = [ds.split("/")[-1]]

# Loop through each fc/table in dictionary
log_it("Looping through each feature/class table")
report_dict = (
    {}
)  # {ds name: {feature count: 0, domain fields: [{field: '', domain type: '', valid_values: '', invalid: [{invalid value: '', count: 0}]}]}
for fds, ds_list in ds_dict.items():
    if fds != "stand_alone":
        log_it(f"Processing feature classes from feature dataset: {fds}")
        # Set workspace path to fds
        arcpy.env.workspace = os.path.join(in_ws, fds)
    else:
        log_it(f"Processing stand-alone feature classes and tables from workspace")
        # Set workspace to gdb
        arcpy.env.workspace = in_ws

    # Loop through fcs in fds
    for ds in ds_list:
        log_it(ds)
        name = fds + "/" + ds
        name = name.replace("stand_alone/", "")
        # Get feature count
        feat_count = arcpy.management.GetCount(ds).getOutput(0)

        # Get fields with domains
        domain_flds = [fld for fld in arcpy.ListFields(ds) if fld.domain]
        # Loop through domain fields to get domain properties
        fld_list = []
        for fld in domain_flds:
            invalid_list = []
            domain_name = fld.domain
            # Look up domain in dictionary
            domain = domain_dict[domain_name]
            # Get domain properties
            domain_type = domain.domainType

            # Get domain values/ranges
            if domain_type == "CodedValue":
                valid_values = tuple(domain.codedValues.keys())
                where = f"{fld.name} NOT IN {valid_values}"
                unique_list = []
                with arcpy.da.SearchCursor(
                    ds, [fld.name], where, sql_clause=("DISTINCT", None)
                ) as cur:
                    for row in cur:
                        unique_list.append(row[0])
                # Update valid values to string so it can be added to excel
                valid_values = ",".join(map(str, valid_values))
                # If there are no invalid values, continue
                if not unique_list:
                    continue
                else:
                    # Get count of each invalid value
                    for val in unique_list:
                        # Handle string vs number
                        if fld.type == "String":
                            where = f"{fld.name} = '{val}'"
                        else:
                            where = f"{fld.name} = {val}"

                        # Handle fc vs table
                        if arcpy.Describe(ds).dataType == "FeatureClass":
                            arcpy.management.MakeFeatureLayer(ds, "i", where)
                        else:
                            arcpy.management.MakeTableView(ds, "i", where)

                        # Get count of invalid values
                        count = arcpy.management.GetCount("i").getOutput(0)
                        invalid_list.append({"value": val, "count": count})
            else:
                min_range = domain.range[0]
                max_range = domain.range[1]
                valid_values = f"{min_range} - {max_range}"

                # Find where value is less than min range
                where = f"{fld.name} < {min_range}"
                # Handle fc vs table
                if arcpy.Describe(ds).dataType == "FeatureClass":
                    arcpy.management.MakeFeatureLayer(ds, "i", where)
                else:
                    arcpy.management.MakeTableView(ds, "i", where)

                # Get count of invalid values
                count = arcpy.management.GetCount("i").getOutput(0)
                if int(count) > 0:
                    invalid_list.append({"value": f"< {min_range}", "count": count})

                # Find where value is greater than min range
                where = f"{fld.name} > {max_range}"
                # Handle fc vs table
                if arcpy.Describe(ds).dataType == "FeatureClass":
                    arcpy.management.MakeFeatureLayer(ds, "i", where)
                else:
                    arcpy.management.MakeTableView(ds, "i", where)

                # Get count of invalid values
                count = arcpy.management.GetCount("i").getOutput(0)
                if int(count) > 0:
                    invalid_list.append({"value": f"> {max_range}", "count": count})

            # Only add info if invalid values were found
            if invalid_list:
                domain_info = {
                    "field": fld.name,
                    "domain_type": domain_type,
                    "valid_values": valid_values,
                    "invalid": invalid_list,
                }
                fld_list.append(domain_info)

        # Only add info if fields with invalid values were found
        if fld_list:
            # Sort fields alphabetically
            sorted_fld_list = sorted(fld_list, key=lambda x: x["field"])
            report_dict[name] = {
                "feature_count": feat_count,
                "domain_fields": sorted_fld_list,
            }

if report_dict:
    # Create new workbook
    wb = openpyxl.Workbook()
    wb.remove(wb["Sheet"])

    # Loop through report dictionary and write data to excel file
    for ds, values in report_dict.items():
        ws = wb.create_sheet(ds.split("/")[-1])
        bold_font = openpyxl.styles.Font(bold=True)
        ws["A1"] = ds
        ws["A1"].font = bold_font
        ws["A2"] = "Feature Count"
        ws["A2"].font = bold_font
        ws["B2"] = int(values["feature_count"])
        ws["B2"].font = bold_font
        ws["A4"] = "Field Name"
        ws["A4"].font = bold_font
        ws["B4"] = "Domain Type"
        ws["B4"].font = bold_font
        ws["C4"] = "Valid Values"
        ws["C4"].font = bold_font
        ws["D4"] = "Invalid Value"
        ws["D4"].font = bold_font
        ws["E4"] = "Count"
        ws["E4"].font = bold_font

        row = 5
        for domain_fld in values["domain_fields"]:
            ws[f"A{row}"] = domain_fld["field"]
            ws[f"B{row}"] = domain_fld["domain_type"]
            ws[f"C{row}"] = domain_fld["valid_values"]
            merge_start = row
            row += 1
            for val in domain_fld["invalid"]:
                ws[f"D{row}"] = val["value"]
                ws[f"E{row}"] = int(val["count"])
                merge_end = row
                row += 1

            # Merge invalid values cells in columns A, B, and C
            ws.merge_cells(f"A{merge_start}:A{merge_end}")
            ws.merge_cells(f"B{merge_start}:B{merge_end}")
            ws.merge_cells(f"C{merge_start}:C{merge_end}")

        # Update formatting for record count columns
        ws["B2"].number_format = "#,##0"
        for cell in ws["E"]:
            cell.number_format = "#,##0"

        # Center count data in column E
        for row in ws.iter_rows(min_row=1, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set width for column C
        ws.column_dimensions["C"].width = 40

        # Word wrap column with valid vaues
        for cell in ws["C"]:
            cell.alignment = Alignment(wrapText=True)

        # Center count data in columns A, B, and C
        for row in ws.iter_rows(min_row=5, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(vertical="center")

        # Apply autofit to all columns
        autofit_column_widths(ws)

    # Save excel file
    wb.save(out_xls)
    # Open excel file
    os.startfile(out_xls)
