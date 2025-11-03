# ----------------------------------------------------------------------------------------------------------------------------
# List Domain Codes and Ranges and Field Values.py
# Description: Script to loop through all the feature classes in the workspace and report domain codes/ranges and field values
# ----------------------------------------------------------------------------------------------------------------------------

# Import arcpy module
import arcpy, string, os, datetime, time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from os.path import basename
start = time.time()

# Local variables:
ws = arcpy.GetParameterAsText(0)
outFile = arcpy.GetParameterAsText(1)
if '.xlsx' not in outFile:
    outFile = outFile + '.xlsx'

# Create Workbook
wBook = Workbook()

# loop through all the feature classes and tables in the workspace
ctr = 1

domains = arcpy.da.ListDomains(ws)
walk = arcpy.da.Walk(ws, datatype=["FeatureClass", "Table"])
for dirpath, dirname, filenames in walk:
    for filename in filenames:
        item = os.path.join(dirpath, filename)
        record_count = arcpy.GetCount_management(item)
        desc = arcpy.Describe(item)
        splitList = desc.baseName.split('.')
        fcName = splitList[-1]
        arcpy.AddMessage("Working on " + fcName)

        # Create Worksheet
        wBook.create_sheet(fcName)
        wSheet = wBook[fcName]
        # If feature class is empty, change the color of the Sheet to Red
        result = arcpy.GetCount_management(item)
        count = int(result.getOutput(0))
        if count == 0:
            wSheet.sheet_properties.tabColor = "FF0000"

        subtype_field_name = desc.subtypeFieldName
        if not subtype_field_name:
            arcpy.AddMessage(fcName + " does not have a subtype field.")
            # Write header
            wSheet["A1"] = "Feature Class or Table"
            wSheet["B1"] = "Field Name"
            wSheet["C1"] = "Domain Name"
            wSheet["D1"] = "Domain Codes or Range"
            wSheet["E1"] = "Field Values"
            flds = arcpy.ListFields(item)
            rowCtr = 2
            for fld in flds:
                if fld.domain != "":
                    for domain in domains:
                        if domain.name == fld.domain and domain.domainType == "CodedValue":
                            codes = ""
                            for code, desc in domain.codedValues.items():
                                if fld.type == "String":
                                    codes = codes + "," + code
                                if fld.type in ("Double", "Integer", "SmallInteger", "Long"):
                                    codes = codes + "," + str(code)

                        if domain.name == fld.domain and domain.domainType == "Range":
                            codes = ""
                            min_range = domain.range[0]
                            max_range = domain.range[1]
                            codes = codes + "," + str(min_range) + "-" + str(max_range)
                    codes = codes[1:]

                    fldValues = ""
                    arcpy.analysis.Frequency(item, "in_memory" + "\\tbl" + str(ctr), fld.name)
                    with arcpy.da.SearchCursor("in_memory" + "\\tbl" + str(ctr), fld.name) as cursor:
                        for row in cursor:
                            fldValue = row[0]
                            if fldValue == None:
                                fldValue = "Null"
                            if fld.type == "String":
                                fldValues = fldValues + "," + fldValue
                            if fld.type in ("Double", "Integer", "SmallInteger"):
                                fldValues = fldValues + "," + str(fldValue)
                    ctr += 1
                    fldValues = fldValues[1:]

                    wSheet["A" + str(rowCtr)] = fcName
                    wSheet["B" + str(rowCtr)] = fld.name
                    wSheet["C" + str(rowCtr)] = fld.domain
                    wSheet["D" + str(rowCtr)] = codes
                    wSheet["E" + str(rowCtr)] = fldValues
                    rowCtr += 1

        else:
            arcpy.AddMessage(fcName + " feature class has a subtype field.")
            # Write header
            wSheet["A1"] = "Feature Class or Table"
            wSheet["B1"] = "Subtype Code"
            wSheet["C1"] = "Subtype Description"
            wSheet["D1"] = "Field Name"
            wSheet["E1"] = "Domain Name"
            wSheet["F1"] = "Domain Codes or Range"
            wSheet["G1"] = "Field Values"
            subtypes = arcpy.da.ListSubtypes(item)
            rowCtr = 2

            for stcode, stdict in list(subtypes.items()):
                for stkey in list(stdict.keys()):
                    if stkey == "Name":
                        sTypeCode = stdict["Name"]
                    if stkey == "FieldValues":
                        fields = stdict[stkey]
                        for field, fieldvals in list(fields.items()):
                            if fieldvals[1] is not None:
                                for domain in domains:
                                    if domain.name == fieldvals[1].name and fieldvals[1].domainType == "CodedValue":
                                        domainName = domain.name
                                        codes = ""
                                        for code, desc in domain.codedValues.items():
                                            if fieldvals[1].type == "Text":
                                                codes = codes + "," + code
                                            if fieldvals[1].type in ("Double", "Short", "Long"):
                                                codes = codes + "," + str(code)

                                    if domain.name == fieldvals[1].name and fieldvals[1].domainType == "Range":
                                        domainName = domain.name
                                        codes = ""
                                        min_range = domain.range[0]
                                        max_range = domain.range[1]
                                        codes = codes + "," + str(min_range) + "-" + str(max_range)
                                codes = codes[1:]

                                fldValues = ""
                                arcpy.analysis.Frequency(item, "in_memory" + "\\tbl" + str(ctr), field)
                                with arcpy.da.SearchCursor("in_memory" + "\\tbl" + str(ctr), field) as cursor:
                                    for row in cursor:
                                        fldValue = row[0]
                                        if fldValue == None:
                                            fldValue = "Null"
                                        if fld.type == "String":
                                            fldValues = fldValues + "," + str(fldValue)
                                        if fld.type in ("Double", "Integer", "SmallInteger", "Long"):
                                            fldValues = fldValues + "," + str(fldValue)
                                fldValues = fldValues[1:]
                                ctr += 1

                                wSheet["A" + str(rowCtr)] = fcName
                                wSheet["B" + str(rowCtr)] = stcode
                                wSheet["C" + str(rowCtr)] = sTypeCode
                                wSheet["D" + str(rowCtr)] = field
                                wSheet["E" + str(rowCtr)] = domainName
                                wSheet["F" + str(rowCtr)] = codes
                                wSheet["G" + str(rowCtr)] = fldValues
                                rowCtr += 1



        # Freeze the first column and row
        c = wSheet['B2']
        wSheet.freeze_panes = c
        # set first row to be Bold
        bold_font = Font(bold=True)
        for cell in wSheet["1:1"]:
            cell.font = bold_font

        # Adjust the width of each column
        for col in wSheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            wSheet.column_dimensions[column].width = adjusted_width


# Delete the first sheet named Sheet
if 'Sheet' in wBook.sheetnames:
    sheet_to_delete = wBook['Sheet']
    wBook.remove(sheet_to_delete)
wBook.save(outFile)

end = time.time()
elapsed = int((end-start)/60)
arcpy.AddMessage("Process completed in " + str(elapsed) + " minute(s)")
