"""
Brooke Reams - breams@esri.com
Jan. 5, 2026

Description:
Loops through coded value domains in an input geodatabse
and creates an output Excel report for mapping field values
to their respective domains.  Designed to be used with FME.


ArcGIS Pro 3.5.2
Python 3.11.11

Updates:


"""

import arcpy
import os
import numbers
import re
import difflib
import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.utils import get_column_letter

# Overwrite existing output
arcpy.env.overwriteOutput = 1


def log_it(message):
    print(message)
    arcpy.AddMessage(message)


def update_sheet_name(sheet_name):
    # Check if sheet name ends with a number
    pattern = r"_\d+$"
    m = re.search(pattern, sheet_name)
    if m:
        num = m.group().replace("_", "")
        char_len = len(num)
        return sheet_name[:-1*(char_len)] + f"{int(num)+1}"
    else:
        return sheet_name + "_1"


def get_close_matches(value, possibilities, n=3, cutoff=0.5):
    matches = []

    # Check if attribute value is a substring of code/description
    for p in possibilities:
        # Ignore case
        m = re.search(value, p, re.IGNORECASE)
        if m:
            matches.append(value)
            break

    # Check if any words in value match code/description
    for p in possibilities:
        # Ignore case
        m = re.search(p, value, re.IGNORECASE)
        if m:
            matches.append(value)
            break

    # Use difflib 
    matches.extend(difflib.get_close_matches(value, possibilities, n, cutoff))

    return list(set(matches))


def is_number(value):
    return isinstance(value, numbers.Number) and not isinstance(value, bool)


def autofit_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(
            col[0].column
        )  # Get column letter from the first cell in the column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (
                TypeError
            ):  # Handle cases where cell.value might be None or not easily convertible to string
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column].width = adjusted_width


# Tool inputs
in_ws = arcpy.GetParameterAsText(0) or r"C:\Users\broo4976\OneDrive - Esri\Projects\MichelleJohnson\ReportingTools\Data_Readiness.gdb"
out_xls = arcpy.GetParameterAsText(1) or r"C:\Users\broo4976\OneDrive - Esri\Projects\MichelleJohnson\ReportingTools\test1.xlsx"

# Set workspace environment
arcpy.env.workspace = in_ws
# Get list of feature datasets
fds_list = arcpy.ListDatasets(feature_type="Feature")
fds_list.append("")
# Create field domain dictionary {domain name: [fc:fld, fc:fld]}
domain_fld_dict = {}
# Create dictionary to store unqiue values associated with fc:field {'fc:field': [values]}
attr_values_dict = {}
# Loop through fcs in feature datasets to populate field domain dictionary
for fds in fds_list:
    fc_list = arcpy.ListFeatureClasses(feature_dataset=fds)
    if fds == "":
        # Get tables
        fc_list.extend(arcpy.ListTables())
    for fc in fc_list:
        # Get list of fields with domain assigned
        fld_list = [fld for fld in arcpy.ListFields(fc) if fld.domain]
        for fld in fld_list:
            if fld.domain not in domain_fld_dict.keys():
                domain_fld_dict[fld.domain] = [f"{fc}:{fld.name}"]
            else:
                domain_fld_dict[fld.domain].append(f"{fc}:{fld.name}")

            # Get a list of unique values in field with domain
            attr_values = []
            with arcpy.da.SearchCursor(fc, [fld.name], f"{fld.name} IS NOT NULL", sql_clause=("DISTINCT", None)) as cur:
                for row in cur:
                    attr_values.append(row[0])
            attr_values_dict[f"{fc}:{fld.name}"] = attr_values



# Get a list of coded value domains in workspace
domain_list = [domain for domain in arcpy.da.ListDomains(in_ws) if domain.domainType == "CodedValue"]

# Loop through domains in workspace and get required details for each
report_dict = {} # {domain name: {fields: ['fc:fld', 'fc:fld'], codes: {code: {desc: description, matches: {'fc:fld': match}, close: {'fc:fld': [close match, close match]}}}, others: {'fc:fld': [value, value]}}
for domain in domain_list:
    # Get coded values
    coded_values = domain.codedValues
    # Get list of domain codes
    codes_list = list(coded_values.keys())
    # Get list of domain descriptions
    desc_list = list(coded_values.values())
    # Initialize domain fields list
    domain_fld_list = []
    # Initialize dictionary to store each code, description, and values
    codes_dict = {} # {domain code: {desc: code description, matches: {'fc:fld': match}, close: {'fc:fld': [close match, close match]}}}
    others_dict = {} # {'fc:fld': [value, value, value]
    # Initialize codes dictionary
    for code, desc in coded_values.items():
        codes_dict[code] = {"desc": desc, "matches": {}, "close": {}}
    
    
    # Get fields assigned to current domain
    if domain.name in domain_fld_dict.keys():
        domain_fld_list = domain_fld_dict[domain.name]
        # Loop through each fc:field and get list of unique attribute values
        for domain_fld in domain_fld_list:
            values_list = attr_values_dict[domain_fld]
            # Create list for values used
            used_list = []
            # Check for values that match codes
            for val in values_list:
                if val in codes_list and val not in used_list:
                    # Add value to code's dictionary
                    codes_dict[val]["matches"][domain_fld] = val
                    # Append value to used list
                    used_list.append(val)

            # Get list of unique values in fc:field that didn't match a domain code
            non_match_list = list(set(values_list) - set(used_list))
            used_list2 = []
            for val in non_match_list:
                # Only check for close matches if value is not a number or less than 3 char
                if not is_number(val) and len(val) > 2:
                    for code, desc in coded_values.items():
                        if domain_fld in codes_dict[code]["close"].keys():
                            # Get current close matches and add to it
                            close_matches = codes_dict[code]["close"][domain_fld]
                            close_matches.extend(get_close_matches(val, [code, desc]))
                        else:
                            # Create list of close matches
                            close_matches = get_close_matches(val, [code, desc])
                        # Close matches function could return actual code as a 'close match'
                        if code in close_matches:
                            close_matches.append(val)
                        # Remove any actual codes from close matches since these are an exact match
                        close_matches = list(set(close_matches) - set(codes_list))
                        if desc in close_matches:
                            # If the value in the field matches the description of a domain,
                            # then remove all close matches other than description because this
                            # is likely the domain that should be used
                            close_matches = [desc]
                        # Remove any duplicate values from the list of close matches
                        codes_dict[code]["close"][domain_fld] = list(set(close_matches))
                        for i in close_matches:
                            # Append close matches to a list so that they are not re-evaluated
                            # as 'other'
                            used_list2.append(i)

            # Check for other values that are not matches and are not close matches
            others_list = list(set(non_match_list) - set(used_list2))
            others_dict[domain_fld] = others_list
            
    # Add data to domain dictionary
    report_dict[domain.name] = {"fields": domain_fld_list, "codes": codes_dict, "others": others_dict}


# Loop through domain dictionary and print info to Excel
if report_dict:
    # Create new workbook
    wb = openpyxl.Workbook()
    wb.remove(wb["Sheet"])

    #  Loop through report dictionary and write data to excel file
    sheet_names = []
    for domain_name, domain_details in report_dict.items():
        # Sheet name has char limit of 31 chars
        sheet_name = domain_name[:31]
        if sheet_name in sheet_names:
            # Sheet name already exists - make sheet name unique
            sheet_name = update_sheet_name(domain_name[:31])
        # Add sheet name to list of sheet names
        sheet_names.append(sheet_name)
        # Create sheet
        ws = wb.create_sheet(sheet_name)
        domain_fld_list = domain_details["fields"]
        # Add column headers
        ws["A1"] = "Code"
        ws["B1"] = "Description"
        col = 3
        for fld in domain_fld_list:
            ws.cell(row=1, column=col, value=fld)
            col+=1       

        # Bold first row
        bold_font = openpyxl.styles.Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font

        # Create yellow and orange fills for hightlighted cells
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FF991C", end_color="FF991C", fill_type="solid")

        # Add coded value details
        codes_dict = domain_details["codes"]
        # Start adding data to row 2, after header cells
        row = 2
        for code, code_details in codes_dict.items():
            matches = code_details["matches"]
            # If the code does not have any matching values, just add code and desc
            if not matches:
                ws[f"A{row}"] = code
                ws[f"B{row}"] = code_details["desc"]
            for fld, val in matches.items():
                ws[f"A{row}"] = code
                ws[f"B{row}"] = code_details["desc"]
                # Get column by getting index position in list of fields
                # then add 3 to account for list index beginning at 0, and
                # to skip the code and description columns
                col = domain_fld_list.index(fld) + 3
                ws.cell(row=row, column=col, value=val)
            close_matches = code_details["close"]
            close_start_row = row
            for fld, val_list in close_matches.items():
                row = close_start_row
                for val in val_list:
                    row+=1
                    ws[f"A{row}"] = code
                    ws[f"B{row}"] = code_details["desc"]
                    # Get column by getting index position in list of fields
                    # then add 3 to account for list index beginning at 0, and
                    # to skip the code and description columns
                    col = domain_fld_list.index(fld) + 3
                    highlight_cell = ws.cell(row=row, column=col, value=val)
                    highlight_cell.fill = yellow_fill

            row+=1

        # Add other values to bottom of field and highlight orange
        others_dict = domain_details["others"]
        others_start_row = row
        for fld, val_list in others_dict.items():
            row = others_start_row
            col = domain_fld_list.index(fld) + 3
            for val in val_list:
                highlight_cell = ws.cell(row=row, column=col, value=val)
                highlight_cell.fill = orange_fill
                row+=1
        
        # Apply autofit to all columns
        autofit_column_widths(ws)

    # Sort sheets alphabeically by name
    wb._sheets.sort(key=lambda ws: ws.title.lower())
    # Reset the active sheet
    wb.active = 0
    
    # Save excel
    wb.save(out_xls)

    # Start file
    os.startfile(out_xls)
