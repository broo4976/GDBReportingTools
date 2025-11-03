"""
Brooke Reams - breams@esri.com
Oct. 21, 2025

Description:
Loops through all feature classes and tables at the root
and feature dataset level and reports field properties
on each dataset.

ArcGIS Pro 3.5.2
Python 3.11.11

Updates:

"""

import arcpy
import os
import openpyxl
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule
##import warnings

# Overwrite existing output
arcpy.env.overwriteOutput = 1

# Suppress all UserWarnings from the openpyxl module
##warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def log_it(message):
    print(message)
    arcpy.AddMessage(message)


def autofit_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column) # Get column letter from the first cell in the column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError: # Handle cases where cell.value might be None or not easily convertible to string
                pass
        adjusted_width = (max_length * 1.05) # Add some padding
        ws.column_dimensions[column].width = adjusted_width


# Tool inputs
in_ws = arcpy.GetParameterAsText(0)
out_xls = arcpy.GetParameterAsText(1)

# Create new workbook and define header
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Set workspace environment
arcpy.env.workspace = in_ws

# Initialize list to store data properties
records = []

# Loop through all feature classes/tables in gdb and in all feature datasets
fds_list = arcpy.ListDatasets(feature_type="Feature")
fds_list.sort()
fds_list.append("")
for fds in fds_list:
    # Get feature classes/tables
    if fds == "":
        log_it(f"Processing stand-alone datasets")
        fds = "<standalone>"
        fc_list = [fc for fc in arcpy.ListFeatureClasses() if not fc.lower().startswith("gdb_")]
        fc_list.sort()
        t_list = [t for t in arcpy.ListTables() if not t.lower().startswith("gdb_")]
        t_list.sort()
        ds_list = fc_list + t_list
    else:
        log_it(f"Processing feature dataset: {fds}")
        ds_list = [fc for fc in arcpy.ListFeatureClasses(feature_dataset=fds) if not fc.lower().startswith("gdb_")]
        ds_list.sort()
        
    for ds in ds_list:
        log_it(f"Processing dataset: {ds}")
        # Create new worksheet
        desc = arcpy.Describe(ds)
        if desc.dataType == "FeatureClass":
            tab_name = f"FC_{ds}"
        else:
            tab_name = f"T_{ds}"
        ws = wb.create_sheet(tab_name[:31])
        # Define header
        ws["A1"] = "Feature Dataset"
        ws["B1"] = "Feature Class/Table"
        ws["C1"] = "Field Name"
        ws["D1"] = "Field Type"
        ws["E1"] = "Field Length"
        ws["F1"] = "Default Domain"
        ws["G1"] = "Rows Filled"
        ws["H1"] = "Fill Factor"

        # Initialize start row
        row = 2
        
        # Get record count
        record_count = int(arcpy.management.GetCount(ds).getOutput(0))
        # Get field info
        flds_list = [fld for fld in arcpy.ListFields(ds)]
        for fld in flds_list:
            if fld.type == "":
                where_clause = f"{fld.name} IS NOT NULL OR {fld.name} <> '' OR {fld.name} <> ' '"
            else:
                where_clause = f"{fld.name} IS NOT NULL"
            # Query data to get field counts
            arcpy.management.MakeTableView(ds, "tv", where_clause)
            fld_count = int(arcpy.management.GetCount("tv").getOutput(0))
            if record_count > 0:
                perc = float(fld_count/record_count)
            else:
                perc = 0

            # Populate current worksheet
            ws.cell(row=row, column=1, value=fds)
            ws.cell(row=row, column=2, value=ds)
            ws.cell(row=row, column=3, value=fld.name)
            ws.cell(row=row, column=4, value=fld.type)
            ws.cell(row=row, column=5, value=fld.length)
            ws.cell(row=row, column=6, value=fld.domain)
            ws.cell(row=row, column=7, value=fld_count)
            ws.cell(row=row, column=8, value=perc)
            
            # Update row
            row+=1

        # Bold and freeze first row
        bold_font = openpyxl.styles.Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font
        ws.freeze_panes = "A2"

        # Update formatting for rows filled column
        for cell in ws["G"]:
            cell.number_format = "#,##0"

        # Apply autofit to all columns
        autofit_column_widths(ws)

        # Format percentage row
        for cell in ws["H"]:
            if isinstance(cell.value, (int, float)):  # Only format numeric cells
                cell.number_format = FORMAT_PERCENTAGE_00

        # Conditional formatting on the percentage field
        rule = ColorScaleRule(start_type="percentile", start_value=10, start_color="f8696b",
        mid_type="percentile", mid_value=50, mid_color="FFEF9C",
        end_type="percentile", end_value=90, end_color="63BE7B")
        ws.conditional_formatting.add(f"H2:H{row}", rule)

# Sort sheets alphabetically
wb._sheets.sort(key=lambda ws: ws.title)
            
# Save excel
wb.save(out_xls)

# Start file
os.startfile(out_xls)
            
        
